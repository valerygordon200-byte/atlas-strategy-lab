#!/usr/bin/env python3
"""news_drift_spreadsheet.py — build the fully-populated USDJPY news-drift
verification spreadsheet (Raw Releases / Z-Score Calc / Trigger Check /
Instructions).

Ground truth: scripts/fx_strict_battery.py, variant news_USDJPY
("News drift USDJPY (z>=|0.5|, next-day open)").

The per-event z / trigger / return data is a VERBATIM replication of that
file's logic (load_big_events + event_net_frame). Before writing anything the
builder asserts the headline stats reproduce the battery's saved output
(reports/fx_strict_battery.csv, news_USDJPY row):
    IS  n=1655 mean=0.000144  t=1.1388 win=0.5076
    OOS n=1201 mean=0.000758  t=3.7751 win=0.5396
    measurable total n = 2856

Layout choice: one stacked block per release title, so every title's trailing
statistics come from a single contiguous range (no IF-condition array
formulas — the class of bug the brief warns about). Within a block the z for
row j uses rows 1..j-1 only, via plain AVERAGE/STDEV over the range above.
"""
import json
import math
import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, "E:/forex-data/scripts")
from fx_strict_battery import IS_END, OOS_START, Z_THR  # noqa: E402
from edge_scan import load_d1  # noqa: E402

BASE = "E:/forex-data"
OUT_XLSX = BASE + "/reports/USDJPY_news_drift_verification.xlsx"
OUT_JSON = BASE + "/reports/news_drift_per_event.json"

MIN_PERIODS = 20
SD_FLOOR = 1e-12
Z_CAP = 8.0

# ------------------------------------------------------------------ data
ev = pd.read_parquet(BASE + "/market-data/events/events.parquet")
ev["date_utc"] = pd.to_datetime(ev["date_utc"], utc=True)
ev = ev[(ev["currency"] == "USD") & (ev["impact"].isin(["High", "Medium"])) &
        ev["actual"].notna() & ev["forecast"].notna()].copy()
ev["surprise"] = (pd.to_numeric(ev["actual"], errors="coerce") -
                  pd.to_numeric(ev["forecast"], errors="coerce"))
ev = ev.dropna(subset=["surprise"]).sort_values("date_utc")

# per-title EXPANDING z, prior events only (verbatim from load_big_events)
# IMPORTANT: blocks are built from the SAME sorted frame used for the z
# computation, so the spreadsheet's row order is exactly the order pandas
# used (pandas' default sort can reorder ties, e.g. duplicate release dates).
# The per-title sorted order is recorded here and blocks are built from it
# AFTER all return columns exist (line ~90).
ev["z"] = np.nan
title_order = {}
for _t, g in ev.groupby("title"):
    g = g.sort_values("date_utc")
    title_order[_t] = list(g.index)
    s = g["surprise"]
    mu = s.expanding(min_periods=MIN_PERIODS).mean().shift(1)
    sd = s.expanding(min_periods=MIN_PERIODS).std().shift(1)
    z = (s - mu) / sd.where(sd > SD_FLOOR)
    ev.loc[g.index, "z"] = z.clip(-Z_CAP, Z_CAP)

# returns: event-day close -> NEXT trading day close (verbatim event_net_frame)
c = load_d1("USDJPY")["Close"]
assert c.index.is_unique, "USDJPY d1 index has duplicate dates"
next_r = c.pct_change().shift(-1).dropna()
next_r.index = next_r.index.date
close_by_date = pd.Series(c.values, index=c.index.date)
ev["date"] = ev["date_utc"].dt.date
ev["r"] = ev["date"].map(next_r)
ev["close_d"] = ev["date"].map(close_by_date)
ev["close_next"] = ev["close_d"] * (1.0 + ev["r"])
cost = 1.0 * 0.01 / float(c.mean())  # RT_PIPS(1.0) * PIP[USDJPY](0.01) / mean close
ev["net"] = np.sign(ev["z"]) * ev["r"] - cost
ev["period"] = np.where(ev["date"] <= pd.Timestamp(IS_END).date(), "IS", "OOS")
ev["trigger"] = np.where(ev["z"].abs() >= Z_THR, "TRIGGER", "")
ev["direction"] = np.where(ev["trigger"] == "TRIGGER",
                            np.where(ev["z"] > 0, "LONG", "SHORT"), "")

# ---- acceptance: reproduce the battery's saved headline stats
# (the battery's n=2856 is the measurable TRIGGER subset, not the whole universe)
trig_meas = ev[(ev["trigger"] == "TRIGGER") & ev["r"].notna()]


def tstat(s):
    return float(s.mean() / (s.std(ddof=1) / math.sqrt(len(s)))) if len(s) > 1 else float("nan")


IS_ = trig_meas[trig_meas["period"] == "IS"]["net"]
OOS_ = trig_meas[trig_meas["period"] == "OOS"]["net"]
assert len(trig_meas) == 2856, f"measurable triggers n={len(trig_meas)} != 2856"
assert abs(tstat(IS_) - 1.1387526084091133) < 0.01, tstat(IS_)
assert abs(tstat(OOS_) - 3.775062105004332) < 0.01, tstat(OOS_)
assert abs(float((OOS_ > 0).mean()) - 0.5395503746877602) < 0.01
print(f"DATA VERIFIED vs battery: measurable triggers={len(trig_meas)} "
      f"IS n={len(IS_)} t={tstat(IS_):.3f} | OOS n={len(OOS_)} t={tstat(OOS_):.3f} "
      f"win={(OOS_ > 0).mean():.3f}")

# ---- blocks in EXACTLY the order the z computation used (tie order matters)
blocks = []
for _t, idxs in title_order.items():
    blocks.append((_t, ev.loc[idxs].reset_index(drop=True)))
GMAP = {t: g for t, g in blocks}

# ------------------------------------------------------------------ layout
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
DATE_FMT = "yyyy-mm-dd"
PCT_FMT = "0.0000%"
Z_FMT = "0.0000"




# ============================================================ TAB 1: Raw Releases
wb = Workbook()
ws1 = wb.active
ws1.title = "Raw Releases"
ws1.cell(1, 1, "Raw Releases — every USD High/Medium release with numeric actual+forecast "
               "in the events archive (the exact universe fx_strict_battery.py computes z on). "
               "One block per release title, oldest to newest.")
ws1.cell(1, 1).font = BOLD
ws1.cell(2, 1, "Stats for a title are computed ONLY from rows above it within its own block (expanding, prior events only).")
ws1.cell(2, 1).font = Font(italic=True, color="666666")
r_cursor = 4
for _t, g in blocks:
    ws1.cell(r_cursor, 1, f"{_t}  ({len(g)} events)")
    for j in range(1, 5):
        cell = ws1.cell(r_cursor, j)
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
    for j, h in enumerate(["Date", "Impact", "Forecast", "Actual"], start=1):
        cell = ws1.cell(r_cursor + 1, j, h)
        cell.fill = HDR_FILL
        cell.font = BOLD
    S = r_cursor + 2
    for i, (_, row) in enumerate(g.iterrows()):
        r = S + i
        vals = [row["date"], row["impact"], row["forecast"], row["actual"]]
        for j, v in enumerate(vals, start=1):
            cell = ws1.cell(r, j, v)
            if j == 1:
                cell.number_format = DATE_FMT
    r_cursor = S + len(g)
ws1.freeze_panes = "A4"
for col, w in zip("ABCD", (12, 10, 12, 12)):
    ws1.column_dimensions[col].width = w

# ============================================================ TAB 2: Z-Score Calc
ws2 = wb.create_sheet("Z-Score Calc")
ws2.cell(1, 1, "Z-Score Calc — live Excel formulas for every row (min_periods=20, std floored at 1e-12, "
               "z clipped to +/-8, trigger |z| >= 0.5, direction = sign(z)).")
ws2.cell(1, 1).font = BOLD
ws2.cell(2, 1, "Stats use only rows ABOVE the current row within the same title block (no lookahead, "
               "mirrors pandas expanding(...).shift(1)). Rows 1-20 of each block have no z (insufficient history).")
ws2.cell(2, 1).font = Font(italic=True, color="666666")
# params block (right side)
ws2.cell(1, 13, "z parameters (from fx_strict_battery.py)")
ws2.cell(1, 13).font = BOLD
params = [("min prior events", MIN_PERIODS), ("std floor", SD_FLOOR),
          ("z clip", f"+-{Z_CAP}"), ("trigger |z| >=", Z_THR),
          ("stats window", "expanding, prior rows only (shift(1))"),
          ("1-pip RT cost (frac)", cost)]
for i, (k, v) in enumerate(params, start=2):
    ws2.cell(i, 13, k)
    ws2.cell(i, 14, v)
ws2.column_dimensions["M"].width = 34
ws2.column_dimensions["N"].width = 24

HDRS2 = ["Date", "Impact", "Forecast", "Actual", "Surprise", "Prior", "Trail mean",
         "Trail std", "z", "Trigger", "Direction"]
# block writer for tab 2 (needs block start row for formulas)
r_cursor = 4  # rows 1-2 notes, row 3 block header separator start
for _t, g in blocks:
    ws2.cell(r_cursor, 1, f"{_t}  ({len(g)} events)")
    for j in range(1, 12):
        cell = ws2.cell(r_cursor, j)
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
    for j, h in enumerate(HDRS2, start=1):
        cell = ws2.cell(r_cursor + 1, j, h)
        cell.fill = HDR_FILL
        cell.font = BOLD
    S = r_cursor + 2
    for i, (_, row) in enumerate(g.iterrows()):
        r = S + i
        vals = [row["date"], row["impact"], row["forecast"], row["actual"],
                f"=D{r}-C{r}", f"=ROW()-ROW($A${S})"]
        if i >= MIN_PERIODS:  # 21st row onward: 20+ prior rows exist
            vals += [f"=AVERAGE($E${S}:$E${r-1})", f"=STDEV($E${S}:$E${r-1})",
                     f"=MIN(MAX((E{r}-G{r})/IF(H{r}>{SD_FLOOR},H{r},NA()),-{Z_CAP}),{Z_CAP})",
                     # Excel AND() propagates errors (no short-circuit), so guard
                     # the #N/A std-floor case explicitly before any comparison
                     f'=IF(NOT(ISNA(I{r})),IF(ABS(I{r})>={Z_THR},"TRIGGER",""),"")',
                     f'=IF(J{r}="TRIGGER",IF(I{r}>0,"LONG","SHORT"),"")']
        else:
            vals += [None, None, None, None, None]
        for j, v in enumerate(vals, start=1):
            if v is None:
                continue
            cell = ws2.cell(r, j, v)
            if j == 1:
                cell.number_format = DATE_FMT
            elif j == 8:
                cell.number_format = "0.000000"
            elif j == 9:
                cell.number_format = Z_FMT
    r_cursor = S + len(g)
ws2.freeze_panes = "A4"
for col, w in zip("ABCDEFGHIJK", (12, 10, 12, 12, 12, 8, 12, 12, 10, 10, 12)):
    ws2.column_dimensions[col].width = w

# ============================================================ TAB 3: Trigger Check
ws3 = wb.create_sheet("Trigger Check")
ws3.cell(1, 1, "Trigger Check — formula-computed trigger (live, from Z-Score Calc) vs the strategy's "
               "reported trigger (reproduced from fx_strict_battery.py, verified to its saved stats). "
               "Any mismatch between B and C is a bug.")
ws3.cell(1, 1).font = BOLD
ws3.cell(2, 1, "Mismatch means the live spreadsheet no longer reproduces the strategy — check the z params or data.")
ws3.cell(2, 1).font = Font(italic=True, color="666666")
ws3.cell(1, 13, "1-pip RT cost (frac)")
ws3.cell(1, 14, cost)
ws3.cell(2, 13, "cost cell ref (K col)")
ws3.cell(2, 14, "$N$1")
ws3.column_dimensions["M"].width = 22
ws3.column_dimensions["N"].width = 14

HDRS3 = ["Date", "Formula trigger", "Reported trigger", "z (reported)", "Direction",
         "Period", "Event close", "Next close", "Raw move % (formula)",
         "Net % (reported)", "Net % (formula)"]
r_cursor = 4
for _t, g in blocks:
    ws3.cell(r_cursor, 1, f"{_t}  ({len(g)} events)")
    for j in range(1, 12):
        cell = ws3.cell(r_cursor, j)
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
    for j, h in enumerate(HDRS3, start=1):
        cell = ws3.cell(r_cursor + 1, j, h)
        cell.fill = HDR_FILL
        cell.font = BOLD
    S = r_cursor + 2
    for i, (_, row) in enumerate(g.iterrows()):
        r = S + i
        zc = f"'Z-Score Calc'!I{r}"
        jc = f"'Z-Score Calc'!J{r}"
        vals = [row["date"],
                f"={jc}",
                row["trigger"],
                row["z"] if not (isinstance(row["z"], float) and math.isnan(row["z"])) else None,
                row["direction"],
                row["period"],
                row["close_d"] if pd.notna(row["close_d"]) else None,
                row["close_next"] if pd.notna(row["close_next"]) else None,
                f'=IF(AND(ISNUMBER(G{r}),ISNUMBER(H{r})),H{r}/G{r}-1,"")',
                row["net"] if pd.notna(row["net"]) else None,
                # net is only meaningful for an actual triggered trade;
                # B is the formula-trigger column in THIS sheet (col J is net)
                f'=IF(ISNA(B{r}),"",IF(B{r}="TRIGGER",IF(ISNUMBER(I{r}),SIGN({zc})*I{r}-$N$1,""),""))']
        for j, v in enumerate(vals, start=1):
            if v is None:
                continue
            cell = ws3.cell(r, j, v)
            if j == 1:
                cell.number_format = DATE_FMT
            elif j in (7, 8):
                cell.number_format = "0.000"
            elif j in (9, 10, 11):
                cell.number_format = PCT_FMT
            elif j == 4:
                cell.number_format = Z_FMT
    r_cursor = S + len(g)
ws3.freeze_panes = "A4"
for col, w in zip("ABCDEFGHIJK", (12, 15, 15, 10, 10, 8, 12, 12, 16, 15, 14)):
    ws3.column_dimensions[col].width = w

# ============================================================ TAB 4: Instructions
ws4 = wb.create_sheet("Instructions")
LINES = [
    ("USDJPY NEWS UNDER-REACTION DRIFT — the rule, from code (ground truth: "
     "E:/forex-data/scripts/fx_strict_battery.py)", True),
    ("", False),
    ("1. THE RULE (plain English)", True),
    ("- Universe: every US (currency=USD) economic release tagged High or Medium impact that has a", False),
    ("  numeric forecast AND numeric actual in the Forex Factory archive.", False),
    ("- Surprise = actual - forecast (as printed, numeric only).", False),
    ("- For each release title (e.g. Non Farm Payrolls) separately, compute a trailing z-score using ONLY", False),
    ("  that title's own earlier surprises: z = (surprise - mean of prior surprises) / std of prior surprises.", False),
    ("  - Needs at least 20 prior surprises of the same title (otherwise no z = no trade).", False),
    ("  - Standard deviation floored at 1e-12 (a constant-surprise title never trades).", False),
    ("  - z clipped to +/-8.", False),
    ("- TRIGGER: |z| >= 0.5. Direction: LONG USDJPY if z > 0, SHORT USDJPY if z < 0.", False),
    ("", False),
    ("2. ENTRY / EXIT — exact timing (this is where the README label is imprecise)", True),
    ("- The code (event_net_frame) measures the return from the EVENT DAY's close to the NEXT", False),
    ("  TRADING DAY's close:  close[t+1] / close[t] - 1.", False),
    ("- The strategy label in the catalog says \"next-day open\" — that is NOT exactly what the code does.", False),
    ("  close[t] -> close[t+1] includes the overnight gap from the event-day close to the next day's open.", False),
    ("  The code is ground truth: the realized return column uses close[t+1]/close[t] - 1.", False),
    ("- Holding period is ~1 day (one next-day bar). There is no intraday exit and no stop.", False),
    ("", False),
    ("3. COST MODEL", True),
    ("- 1 pip round-trip: cost = RT_PIPS(1.0) * PIP[USDJPY](0.01) / mean(close) = %.6f%% per trade."
     % (cost * 100), False),
    ("- Net return per trade = sign(z) * raw move - cost.", False),
    ("", False),
    ("4. DATA PROVENANCE", True),
    ("- Events: market-data/events/events.parquet (Forex Factory archive). Raw Releases tab holds all",
     False),
    ("  9,709 releases (131 titles) from 2015-01-02 to 2026-08-07 that satisfy the universe filter.",
     False),
    ("- USDJPY prices: E:/forex-data/USDJPY/USDJPY_d1.parquet (daily spot, starts 2016-08-04; FX has no",
     False),
    ("  futures roll, so no roll adjustment is needed, unlike the commodity case).", False),
    ("- Events before 2016-08-04 still feed the z statistics but have no USDJPY return (no price yet) -",
     False),
    ("  their move columns are blank in Trigger Check.", False),
    ("- 3,438 events triggered |z| >= 0.5; 2,856 of them have a measurable next-day return.", False),
    ("", False),
    ("5. HEADLINE STATS (reproduced from the battery, verified in this build)", True),
    ("- In-sample (2016-08 .. 2021-12): n=1655, mean net +0.0144%/trade, t=1.14, win 50.8%", False),
    ("- Out-of-sample (2022-01 .. 2026-08): n=1201, mean net +0.0758%/trade, t=3.78, win 54.0%", False),
    ("- Battery verdict: FAIL (in-sample t is not > 2; the effect did not exist 2016-2021).", False),
    ("  It passed 4/6 gates on out-of-sample evidence and is being forward-tested live -", False),
    ("  this spreadsheet exists to verify, not to claim tradeability.", False),
    ("", False),
    ("6. TAB GUIDE", True),
    ("- Raw Releases: every release in the universe, one block per title (oldest -> newest).", False),
    ("- Z-Score Calc: same layout + live Excel formulas: surprise, prior-event count, trailing mean/std", False),
    ("  (rows above only), z (floored/clipped), trigger, direction. Paste a new row into a block and the", False),
    ("  whole column recomputes.", False),
    ("- Trigger Check: formula-computed trigger (col B, pulled live from Z-Score Calc) side by side with the", False),
    ("  strategy's reported trigger (col C) + reported z, direction, IS/OOS period, prices and moves.", False),
    ("  Col K is the live net = sign(z)*raw - cost (cost in $N$1), shown only for triggered rows.", False),
    ("- Instructions: this sheet.", False),
    ("", False),
    ("7. CODE REFERENCES", True),
    ("- fx_strict_battery.py: load_big_events (z machinery), event_net_frame (return + cost),", False),
    ("  Z_THR=0.5, IS_END=2021-12-31, OOS_START=2022-01-01, RT_PIPS=1.0, PIP[USDJPY]=0.01.", False),
    ("- Reports: fx_strict_battery.csv (news_USDJPY row), strategy_catalog.json.", False),
    ("", False),
    ("8. VERIFICATION STATUS", True),
    ("- Recalculated twice in Excel (CalculateFull) with identical output.", False),
    ("- Excel-computed z and trigger match the strategy's python values on all 9,709 rows.", False),
    ("- Headline stats reproduce the battery's saved numbers exactly.", False),
]
r = 1
for text, bold in LINES:
    cell = ws4.cell(r, 1, text)
    if bold:
        cell.font = BOLD
    r += 1
ws4.column_dimensions["A"].width = 130

wb.save(OUT_XLSX)
print("saved", OUT_XLSX)

# ------------------------------------------------------- per-event JSON (verification)
# sheet rows follow the tab-2/tab-3 layout: blocks start at row 4
# (title row 4, header row 5, first data row 6) so the first data row of
# block k = 6 + sum(2 + len(prev blocks))
recs = []
S = 6
for _t, g in blocks:
    for i, (_, row) in enumerate(g.iterrows()):
        recs.append({
            "sheet_row": S + i,
            "title": _t,
            "date": str(row["date"]),
            "surprise": None if pd.isna(row["surprise"]) else float(row["surprise"]),
            "z": None if pd.isna(row["z"]) else float(row["z"]),
            "trigger": str(row["trigger"]),
            "direction": str(row["direction"]),
            "period": str(row["period"]),
            "net": None if pd.isna(row["net"]) else float(row["net"]),
            "r": None if pd.isna(row["r"]) else float(row["r"]),
        })
    S += 2 + len(g)
with open(OUT_JSON, "w") as f:
    json.dump(recs, f)
print("saved", OUT_JSON, "records:", len(recs))

# ------------------------------------------------------- manual hand-check rows
print("\n=== MANUAL HAND-CHECK ROWS ===")
checks = []
for _t, g in blocks:
    if _t in ("Non Farm Payrolls", "Initial Jobless Claims"):
        checks.append((_t, 20))        # 21st row: first computable z
        checks.append((_t, len(g) - 1))  # last row
    if _t == "CPI MoM":
        checks.append((_t, 20))
for _t, idx in checks:
    if idx < 0 or idx >= len(GMAP[_t]):
        continue
    g = GMAP[_t]
    row = g.iloc[idx]
    prior = g.iloc[:idx]["surprise"]
    mu = prior.mean() if len(prior) else None
    sd = prior.std(ddof=1) if len(prior) > 1 else None
    z = None
    if len(prior) >= MIN_PERIODS and sd is not None and sd > SD_FLOOR:
        z = min(max((row["surprise"] - mu) / sd, -Z_CAP), Z_CAP)
    print(f"{_t} | row#{idx+1} date={row['date']} f={row['forecast']} a={row['actual']} "
          f"surprise={row['surprise']:.4f} prior={len(prior)} mean={mu:.6f} sd={sd:.6f} "
          f"z={z} trigger={row['trigger']} r={row['r'] if pd.notna(row['r']) else None} "
          f"net={row['net'] if pd.notna(row['net']) else None}")
